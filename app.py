"""
AI Proctoring System - Main Application
A Flask-based web application for AI-powered exam proctoring
"""

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_mysqldb import MySQL
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from functools import wraps
import os
import re
import base64
import json
import io
from io import BytesIO
from datetime import datetime, timedelta
from config import config

# Initialize Flask app
app = Flask(__name__)
app.config.from_object(config['development'])

# Initialize MySQL
mysql = MySQL(app)

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

# Create upload directories
os.makedirs(app.config.get('UPLOAD_FOLDER', 'static/uploads'), exist_ok=True)
os.makedirs(app.config.get('CAPTURED_FOLDER', 'static/captured'), exist_ok=True)


# =============================================
# User Model
# =============================================
class User(UserMixin):
    def __init__(self, id, username, email, full_name, role, face_encoding=None, profile_image=None):
        self.id = id
        self.username = username
        self.email = email
        self.full_name = full_name
        self.role = role
        self.face_encoding = face_encoding
        self.profile_image = profile_image
    
    def is_admin(self):
        return self.role == 'admin'


@login_manager.user_loader
def load_user(user_id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM users WHERE id = %s", (user_id,))
    user_data = cur.fetchone()
    cur.close()
    
    if user_data:
        return User(
            id=user_data['id'],
            username=user_data['username'],
            email=user_data['email'],
            full_name=user_data['full_name'],
            role=user_data['role'],
            face_encoding=user_data['face_encoding'],
            profile_image=user_data['profile_image']
        )
    return None


# =============================================
# Decorators
# =============================================
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            flash('You need admin privileges to access this page.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


# =============================================
# Public Routes
# =============================================
@app.route('/')
def index():
    """Homepage"""
    return render_template('index.html')


@app.route('/about')
def about():
    """About Us page"""
    return render_template('about.html')


# =============================================
# Authentication Routes
# =============================================
@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login page"""
    if current_user.is_authenticated:
        if current_user.is_admin():
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('user_dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM users WHERE username = %s AND password = %s", (username, password))
        user_data = cur.fetchone()
        cur.close()
        
        if user_data:
            user = User(
                id=user_data['id'],
                username=user_data['username'],
                email=user_data['email'],
                full_name=user_data['full_name'],
                role=user_data['role'],
                face_encoding=user_data['face_encoding'],
                profile_image=user_data['profile_image']
            )
            login_user(user)
            flash(f'Welcome back, {user.full_name}!', 'success')
            
            next_page = request.args.get('next')
            if user.is_admin():
                return redirect(next_page or url_for('admin_dashboard'))
            return redirect(next_page or url_for('user_dashboard'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """User registration page"""
    if current_user.is_authenticated:
        return redirect(url_for('user_dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        full_name = request.form.get('full_name')
        
        # Validation
        if not re.match(r'^[A-Za-z\s]+$', full_name):
            flash('Full name can only contain letters and spaces.', 'error')
            return render_template('register.html')
        
        if ' ' in username:
            flash('Username cannot contain spaces.', 'error')
            return render_template('register.html')
        
        if username[0].isdigit():
            flash('Username cannot start with a number.', 'error')
            return render_template('register.html')
        
        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return render_template('register.html')
        
        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('register.html')
        
        cur = mysql.connection.cursor()
        
        # Check if username exists
        cur.execute("SELECT id FROM users WHERE username = %s", (username,))
        if cur.fetchone():
            flash('Username already exists.', 'error')
            cur.close()
            return render_template('register.html')
        
        # Check if email exists
        cur.execute("SELECT id FROM users WHERE email = %s", (email,))
        if cur.fetchone():
            flash('Email already registered.', 'error')
            cur.close()
            return render_template('register.html')
        
        # Insert new user
        cur.execute(
            "INSERT INTO users (username, email, password, full_name, role) VALUES (%s, %s, %s, %s, 'user')",
            (username, email, password, full_name)
        )
        mysql.connection.commit()
        cur.close()
        
        flash('Registration successful! Please log in.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')


@app.route('/logout')
@login_required
def logout():
    """User logout"""
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))


# =============================================
# User Routes
# =============================================
@app.route('/dashboard')
@login_required
def user_dashboard():
    """User dashboard"""
    if current_user.is_admin():
        return redirect(url_for('admin_dashboard'))
    
    cur = mysql.connection.cursor()
    
    # Get available exams
    cur.execute("SELECT * FROM exams WHERE is_active = TRUE ORDER BY created_at DESC")
    exams = cur.fetchall()
    
    # Get user's exam history
    cur.execute("""
        SELECT es.*, e.title as exam_title 
        FROM exam_sessions es 
        JOIN exams e ON es.exam_id = e.id 
        WHERE es.user_id = %s 
        ORDER BY es.created_at DESC 
        LIMIT 10
    """, (current_user.id,))
    history = cur.fetchall()
    
    # Get recent alerts/warnings
    cur.execute("""
        SELECT * FROM proctoring_logs 
        WHERE user_id = %s AND severity IN ('warning', 'critical')
        ORDER BY created_at DESC 
        LIMIT 5
    """, (current_user.id,))
    alerts = cur.fetchall()
    
    cur.close()
    
    return render_template('user/dashboard.html', exams=exams, history=history, alerts=alerts)


@app.route('/exam/<int:exam_id>')
@login_required
def take_exam(exam_id):
    """Start or continue an exam"""
    if current_user.is_admin():
        flash('Admins cannot take exams.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    cur = mysql.connection.cursor()
    
    # Get exam details
    cur.execute("SELECT * FROM exams WHERE id = %s AND is_active = TRUE", (exam_id,))
    exam = cur.fetchone()
    
    if not exam:
        flash('Exam not found or not available.', 'error')
        cur.close()
        return redirect(url_for('user_dashboard'))
    
    # Block re-entry if a session was already terminated for this exam
    cur.execute("""
        SELECT * FROM exam_sessions
        WHERE user_id = %s AND exam_id = %s AND status = 'terminated'
        ORDER BY created_at DESC LIMIT 1
    """, (current_user.id, exam_id))
    terminated_session = cur.fetchone()

    if terminated_session:
        flash('You cannot retake this exam. Your previous session was terminated due to violations.', 'error')
        cur.close()
        return redirect(url_for('user_dashboard'))

    # Check for existing in-progress session
    cur.execute("""
        SELECT * FROM exam_sessions 
        WHERE user_id = %s AND exam_id = %s AND status = 'in_progress'
    """, (current_user.id, exam_id))
    existing_session = cur.fetchone()
    
    if existing_session:
        session_id = existing_session['id']
    else:
        # Create new exam session
        cur.execute("""
            INSERT INTO exam_sessions (user_id, exam_id, status) 
            VALUES (%s, %s, 'in_progress')
        """, (current_user.id, exam_id))
        mysql.connection.commit()
        session_id = cur.lastrowid
        
        # Log exam start
        cur.execute("""
            INSERT INTO proctoring_logs (session_id, user_id, event_type, severity, description)
            VALUES (%s, %s, 'exam_started', 'info', 'User started the exam')
        """, (session_id, current_user.id))
        mysql.connection.commit()
    
    # Get questions
    cur.execute("SELECT * FROM questions WHERE exam_id = %s ORDER BY id", (exam_id,))
    questions = cur.fetchall()
    
    cur.close()
    
    return render_template('user/exam.html', exam=exam, questions=questions, session_id=session_id)


@app.route('/submit_exam', methods=['POST'])
@login_required
def submit_exam():
    """Submit exam answers"""
    data = request.get_json()
    session_id = data.get('session_id')
    answers = data.get('answers', {})
    
    cur = mysql.connection.cursor()
    
    # Get session and exam info
    cur.execute("""
        SELECT es.*, e.passing_marks 
        FROM exam_sessions es 
        JOIN exams e ON es.exam_id = e.id 
        WHERE es.id = %s AND es.user_id = %s
    """, (session_id, current_user.id))
    session_data = cur.fetchone()
    
    if not session_data:
        return jsonify({'success': False, 'message': 'Invalid session'})
    
    # Calculate score
    total_score = 0
    correct_count = 0
    total_questions = 0
    
    for question_id, selected_option in answers.items():
        cur.execute("SELECT correct_option, marks FROM questions WHERE id = %s", (question_id,))
        question = cur.fetchone()
        
        if question:
            total_questions += 1
            is_correct = question['correct_option'] == selected_option
            
            if is_correct:
                total_score += question['marks']
                correct_count += 1
            
            # Save answer
            cur.execute("""
                INSERT INTO exam_answers (session_id, question_id, selected_option, is_correct)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE selected_option = %s, is_correct = %s
            """, (session_id, question_id, selected_option, is_correct, selected_option, is_correct))
    
    # Update session
    status = 'completed'
    cur.execute("""
        UPDATE exam_sessions 
        SET status = %s, score = %s, total_questions = %s, correct_answers = %s, end_time = NOW()
        WHERE id = %s
    """, (status, total_score, total_questions, correct_count, session_id))
    
    # Log exam completion
    cur.execute("""
        INSERT INTO proctoring_logs (session_id, user_id, event_type, severity, description)
        VALUES (%s, %s, 'exam_completed', 'info', %s)
    """, (session_id, current_user.id, f'Exam completed with score {total_score}'))
    
    mysql.connection.commit()
    cur.close()
    
    return jsonify({
        'success': True,
        'score': total_score,
        'correct': correct_count,
        'total': total_questions,
        'passed': total_score >= session_data['passing_marks']
    })


@app.route('/results/<int:session_id>')
@login_required
def view_results(session_id):
    """View exam results"""
    cur = mysql.connection.cursor()
    
    cur.execute("""
        SELECT es.*, e.title as exam_title, e.total_marks, e.passing_marks
        FROM exam_sessions es
        JOIN exams e ON es.exam_id = e.id
        WHERE es.id = %s AND es.user_id = %s
    """, (session_id, current_user.id))
    result = cur.fetchone()
    
    if not result:
        flash('Results not found.', 'error')
        cur.close()
        return redirect(url_for('user_dashboard'))
    
    # Get detailed answers
    cur.execute("""
        SELECT ea.*, q.question_text, q.option_a, q.option_b, q.option_c, q.option_d, q.correct_option
        FROM exam_answers ea
        JOIN questions q ON ea.question_id = q.id
        WHERE ea.session_id = %s
    """, (session_id,))
    answers = cur.fetchall()
    
    cur.close()
    
    return render_template('user/results.html', result=result, answers=answers)


@app.route('/register_face', methods=['GET', 'POST'])
@login_required
def register_face():
    """Register user's face for proctoring"""
    if request.method == 'POST':
        data = request.get_json()
        image_data = data.get('image')
        
        if image_data:
            try:
                # Decode base64 image
                image_data = image_data.split(',')[1]
                image_bytes = base64.b64decode(image_data)
                
                # Save image
                filename = f"user_{current_user.id}_face.jpg"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                
                with open(filepath, 'wb') as f:
                    f.write(image_bytes)
                
                # Try to encode face (simplified - in production use face_recognition library)
                # For demo, we just store the image path
                cur = mysql.connection.cursor()
                cur.execute("""
                    UPDATE users SET face_encoding = %s, profile_image = %s WHERE id = %s
                """, (filename, filename, current_user.id))
                mysql.connection.commit()
                cur.close()
                
                return jsonify({'success': True, 'message': 'Face registered successfully!'})
            except Exception as e:
                return jsonify({'success': False, 'message': str(e)})
        
        return jsonify({'success': False, 'message': 'No image data received'})
    
    return render_template('user/register_face.html')


# =============================================
# Proctoring API Routes
# =============================================
@app.route('/api/verify_face', methods=['POST'])
@login_required
def verify_face():
    """Verify user's face during exam"""
    data = request.get_json()
    session_id = data.get('session_id')
    image_data = data.get('image')
    
    if not image_data or not session_id:
        return jsonify({'verified': False, 'message': 'Missing data'})
    
    try:
        # In production, compare with stored face encoding using face_recognition library
        # For demo, we simulate face verification
        
        cur = mysql.connection.cursor()
        
        # Check if user has registered face
        if current_user.face_encoding:
            # Simulate verification (in production, use actual face comparison)
            verified = True
            event_type = 'face_detected'
            severity = 'info'
            description = 'Face verified successfully'
        else:
            verified = False
            event_type = 'face_not_detected'
            severity = 'warning'
            description = 'No registered face to compare'
        
        # Log the event
        cur.execute("""
            INSERT INTO proctoring_logs (session_id, user_id, event_type, severity, description)
            VALUES (%s, %s, %s, %s, %s)
        """, (session_id, current_user.id, event_type, severity, description))
        mysql.connection.commit()
        cur.close()
        
        return jsonify({'verified': verified, 'message': description})
    except Exception as e:
        return jsonify({'verified': False, 'message': str(e)})


@app.route('/api/log_event', methods=['POST'])
@login_required
def log_proctoring_event():
    """Log proctoring events (tab switch, window blur, etc.)"""
    data = request.get_json()
    session_id = data.get('session_id')
    event_type = data.get('event_type')
    description = data.get('description', '')
    
    if not session_id or not event_type:
        return jsonify({'success': False})
    
    # Determine severity
    severity_map = {
        'tab_switch': 'warning',
        'window_blur': 'warning',
        'multiple_faces': 'critical',
        'face_not_detected': 'warning',
        'face_mismatch': 'critical',
        'suspicious_movement': 'warning'
    }
    severity = severity_map.get(event_type, 'info')
    
    cur = mysql.connection.cursor()
    
    # Log event
    cur.execute("""
        INSERT INTO proctoring_logs (session_id, user_id, event_type, severity, description)
        VALUES (%s, %s, %s, %s, %s)
    """, (session_id, current_user.id, event_type, severity, description))
    
    # Update warning count
    terminated = False
    if severity in ['warning', 'critical']:
        cur.execute("""
            UPDATE exam_sessions SET warning_count = warning_count + 1 WHERE id = %s
        """, (session_id,))
        
        # Check if max warnings exceeded
        cur.execute("SELECT warning_count FROM exam_sessions WHERE id = %s", (session_id,))
        session_data = cur.fetchone()
        
        if session_data and session_data['warning_count'] >= 3:
            terminated = True
            cur.execute("""
                UPDATE exam_sessions SET status = 'terminated', end_time = NOW() WHERE id = %s
            """, (session_id,))
            cur.execute("""
                INSERT INTO proctoring_logs (session_id, user_id, event_type, severity, description)
                VALUES (%s, %s, 'exam_terminated', 'critical', 'Exam terminated due to excessive warnings')
            """, (session_id, current_user.id))
    
    mysql.connection.commit()
    cur.close()
    
    return jsonify({'success': True, 'terminated': terminated})


# =============================================
# Admin Routes
# =============================================
@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    """Admin dashboard with analytics"""
    cur = mysql.connection.cursor()
    
    # Get statistics
    cur.execute("SELECT COUNT(*) as count FROM users WHERE role = 'user'")
    total_users = cur.fetchone()['count']
    
    cur.execute("SELECT COUNT(*) as count FROM exams")
    total_exams = cur.fetchone()['count']
    
    cur.execute("SELECT COUNT(*) as count FROM exam_sessions")
    total_sessions = cur.fetchone()['count']
    
    cur.execute("SELECT COUNT(*) as count FROM exam_sessions WHERE status = 'in_progress'")
    active_sessions = cur.fetchone()['count']
    
    cur.execute("SELECT COUNT(*) as count FROM proctoring_logs WHERE severity = 'critical'")
    critical_incidents = cur.fetchone()['count']
    
    cur.execute("""
        SELECT COUNT(*) as count FROM exam_sessions 
        WHERE status = 'completed' AND score >= (
            SELECT passing_marks FROM exams WHERE id = exam_sessions.exam_id
        )
    """)
    passed_exams = cur.fetchone()['count']
    
    # Recent activity
    cur.execute("""
        SELECT pl.*, u.username, u.full_name
        FROM proctoring_logs pl
        JOIN users u ON pl.user_id = u.id
        ORDER BY pl.created_at DESC
        LIMIT 10
    """)
    recent_activity = cur.fetchall()
    
    # Exam statistics
    cur.execute("""
        SELECT e.title, 
               COUNT(es.id) as attempts,
               AVG(es.score) as avg_score,
               SUM(CASE WHEN es.status = 'flagged' OR es.status = 'terminated' THEN 1 ELSE 0 END) as flagged
        FROM exams e
        LEFT JOIN exam_sessions es ON e.id = es.exam_id
        GROUP BY e.id, e.title
    """)
    exam_stats = cur.fetchall()
    
    cur.close()
    
    stats = {
        'total_users': total_users,
        'total_exams': total_exams,
        'total_sessions': total_sessions,
        'active_sessions': active_sessions,
        'critical_incidents': critical_incidents,
        'passed_exams': passed_exams
    }
    
    return render_template('admin/dashboard.html', stats=stats, recent_activity=recent_activity, exam_stats=exam_stats)


@app.route('/admin/monitor')
@login_required
@admin_required
def admin_monitor():
    """Live monitoring of ongoing exams"""
    cur = mysql.connection.cursor()
    
    # Get active exam sessions
    cur.execute("""
        SELECT es.*, u.username, u.full_name, e.title as exam_title
        FROM exam_sessions es
        JOIN users u ON es.user_id = u.id
        JOIN exams e ON es.exam_id = e.id
        WHERE es.status = 'in_progress'
        ORDER BY es.start_time DESC
    """)
    active_sessions = cur.fetchall()
    
    cur.close()
    
    return render_template('admin/monitor.html', active_sessions=active_sessions)


@app.route('/admin/reports')
@login_required
@admin_required
def admin_reports():
    """Analytics and reports page"""
    cur = mysql.connection.cursor()
    
    # Daily exam statistics (last 7 days)
    cur.execute("""
        SELECT DATE(created_at) as date, 
               COUNT(*) as total_exams,
               SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed,
               SUM(CASE WHEN status = 'flagged' OR status = 'terminated' THEN 1 ELSE 0 END) as flagged
        FROM exam_sessions
        WHERE created_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)
        GROUP BY DATE(created_at)
        ORDER BY date
    """)
    daily_stats = cur.fetchall()
    
    # Proctoring incidents by type
    cur.execute("""
        SELECT event_type, COUNT(*) as count
        FROM proctoring_logs
        WHERE severity IN ('warning', 'critical')
        GROUP BY event_type
        ORDER BY count DESC
    """)
    incident_types = cur.fetchall()
    
    # Top performing exams
    cur.execute("""
        SELECT e.title, 
               COUNT(es.id) as attempts,
               AVG(es.score) as avg_score,
               MAX(es.score) as highest_score
        FROM exams e
        LEFT JOIN exam_sessions es ON e.id = es.exam_id AND es.status = 'completed'
        GROUP BY e.id, e.title
        HAVING attempts > 0
        ORDER BY avg_score DESC
    """)
    top_exams = cur.fetchall()
    
    # User performance summary
    cur.execute("""
        SELECT u.full_name, u.username,
               COUNT(es.id) as exams_taken,
               AVG(es.score) as avg_score,
               SUM(es.warning_count) as total_warnings
        FROM users u
        LEFT JOIN exam_sessions es ON u.id = es.user_id
        WHERE u.role = 'user'
        GROUP BY u.id, u.full_name, u.username
        HAVING exams_taken > 0
        ORDER BY avg_score DESC
        LIMIT 10
    """)
    user_performance = cur.fetchall()
    
    cur.close()
    
    return render_template('admin/reports.html', 
                          daily_stats=daily_stats, 
                          incident_types=incident_types,
                          top_exams=top_exams,
                          user_performance=user_performance)


def _get_report_data():
    """Fetch report data for exports"""
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT DATE(created_at) as date, 
               COUNT(*) as total_exams,
               SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed,
               SUM(CASE WHEN status = 'flagged' OR status = 'terminated' THEN 1 ELSE 0 END) as flagged
        FROM exam_sessions
        WHERE created_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)
        GROUP BY DATE(created_at)
        ORDER BY date
    """)
    daily_stats = cur.fetchall()
    cur.execute("""
        SELECT event_type, COUNT(*) as count
        FROM proctoring_logs
        WHERE severity IN ('warning', 'critical')
        GROUP BY event_type
        ORDER BY count DESC
    """)
    incident_types = cur.fetchall()
    cur.execute("""
        SELECT e.title, 
               COUNT(es.id) as attempts,
               AVG(es.score) as avg_score,
               MAX(es.score) as highest_score
        FROM exams e
        LEFT JOIN exam_sessions es ON e.id = es.exam_id AND es.status = 'completed'
        GROUP BY e.id, e.title
        HAVING attempts > 0
        ORDER BY avg_score DESC
    """)
    top_exams = cur.fetchall()
    cur.execute("""
        SELECT u.full_name, u.username,
               COUNT(es.id) as exams_taken,
               AVG(es.score) as avg_score,
               SUM(es.warning_count) as total_warnings
        FROM users u
        LEFT JOIN exam_sessions es ON u.id = es.user_id
        WHERE u.role = 'user'
        GROUP BY u.id, u.full_name, u.username
        HAVING exams_taken > 0
        ORDER BY avg_score DESC
        LIMIT 10
    """)
    user_performance = cur.fetchall()
    cur.close()
    return daily_stats, incident_types, top_exams, user_performance


@app.route('/admin/reports/export/excel')
@login_required
@admin_required
def admin_reports_export_excel():
    """Export report as Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        flash('Excel export requires openpyxl. Run: pip install openpyxl', 'error')
        return redirect(url_for('admin_reports'))

    daily_stats, incident_types, top_exams, user_performance = _get_report_data()
    wb = Workbook()
    
    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = "Analytics Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    row = 3
    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Value")
    for c in range(1, 3):
        ws.cell(row=row, column=c).fill = header_fill
        ws.cell(row=row, column=c).font = header_font
    row += 1
    
    total_exams = sum(s.get('total_exams', 0) or 0 for s in daily_stats)
    completed = sum(s.get('completed', 0) or 0 for s in daily_stats)
    flagged = sum(s.get('flagged', 0) or 0 for s in daily_stats)
    
    ws.cell(row=row, column=1, value="Total Exams (7 days)")
    ws.cell(row=row, column=2, value=total_exams)
    row += 1
    ws.cell(row=row, column=1, value="Completed Exams")
    ws.cell(row=row, column=2, value=completed)
    row += 1
    ws.cell(row=row, column=1, value="Flagged/Terminated")
    ws.cell(row=row, column=2, value=flagged)
    row += 2
    
    # Daily stats
    ws.cell(row=row, column=1, value="Date")
    ws.cell(row=row, column=2, value="Total")
    ws.cell(row=row, column=3, value="Completed")
    ws.cell(row=row, column=4, value="Flagged")
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = header_fill
        ws.cell(row=row, column=c).font = header_font
    row += 1
    for stat in daily_stats:
        d = stat.get('date')
        ws.cell(row=row, column=1, value=str(d) if d else '')
        ws.cell(row=row, column=2, value=stat.get('total_exams', 0) or 0)
        ws.cell(row=row, column=3, value=stat.get('completed', 0) or 0)
        ws.cell(row=row, column=4, value=stat.get('flagged', 0) or 0)
        row += 1
    
    row += 2
    # Incidents
    ws.cell(row=row, column=1, value="Incident Type")
    ws.cell(row=row, column=2, value="Count")
    for c in range(1, 3):
        ws.cell(row=row, column=c).fill = header_fill
        ws.cell(row=row, column=c).font = header_font
    row += 1
    for inc in incident_types:
        ws.cell(row=row, column=1, value=(inc.get('event_type') or '').replace('_', ' ').title())
        ws.cell(row=row, column=2, value=inc.get('count', 0) or 0)
        row += 1
    
    # Top Exams sheet
    ws2 = wb.create_sheet("Top Exams")
    ws2.cell(row=1, column=1, value="Exam")
    ws2.cell(row=1, column=2, value="Attempts")
    ws2.cell(row=1, column=3, value="Avg Score")
    ws2.cell(row=1, column=4, value="Highest")
    for c in range(1, 5):
        ws2.cell(row=1, column=c).fill = header_fill
        ws2.cell(row=1, column=c).font = header_font
    for i, exam in enumerate(top_exams, start=2):
        ws2.cell(row=i, column=1, value=exam.get('title', ''))
        ws2.cell(row=i, column=2, value=exam.get('attempts', 0) or 0)
        ws2.cell(row=i, column=3, value=round(exam.get('avg_score') or 0, 1))
        ws2.cell(row=i, column=4, value=exam.get('highest_score') or 'N/A')
    
    # User Performance sheet
    ws3 = wb.create_sheet("User Performance")
    headers = ["Rank", "User", "Username", "Exams Taken", "Avg Score", "Total Warnings"]
    for c, h in enumerate(headers, start=1):
        ws3.cell(row=1, column=c, value=h).fill = header_fill
        ws3.cell(row=1, column=c).font = header_font
    for i, u in enumerate(user_performance, start=2):
        ws3.cell(row=i, column=1, value=i - 1)
        ws3.cell(row=i, column=2, value=u.get('full_name', ''))
        ws3.cell(row=i, column=3, value=u.get('username', ''))
        ws3.cell(row=i, column=4, value=u.get('exams_taken', 0) or 0)
        ws3.cell(row=i, column=5, value=round(u.get('avg_score') or 0, 1))
        ws3.cell(row=i, column=6, value=u.get('total_warnings') or 0)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"analytics_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=filename)


@app.route('/admin/reports/export/pdf')
@login_required
@admin_required
def admin_reports_export_pdf():
    """Export report as PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        flash('PDF export requires reportlab. Run: pip install reportlab', 'error')
        return redirect(url_for('admin_reports'))

    daily_stats, incident_types, top_exams, user_performance = _get_report_data()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    story = []
    
    title_style = ParagraphStyle(name='Title', parent=styles['Heading1'], fontSize=18, spaceAfter=20)
    story.append(Paragraph("Analytics &amp; Reports", title_style))
    
    total_exams = sum(s.get('total_exams', 0) or 0 for s in daily_stats)
    completed = sum(s.get('completed', 0) or 0 for s in daily_stats)
    flagged = sum(s.get('flagged', 0) or 0 for s in daily_stats)
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Exams (7 days)", str(total_exams)],
        ["Completed Exams", str(completed)],
        ["Flagged/Terminated", str(flagged)],
    ]
    t = Table(summary_data, colWidths=[3*inch, 2*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#334155')),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    story.append(t)
    story.append(Spacer(1, 20))
    
    story.append(Paragraph("Daily Activity (Last 7 Days)", styles['Heading2']))
    daily_data = [["Date", "Total", "Completed", "Flagged"]]
    for stat in daily_stats:
        d = stat.get('date')
        daily_data.append([
            str(d) if d else '',
            str(stat.get('total_exams', 0) or 0),
            str(stat.get('completed', 0) or 0),
            str(stat.get('flagged', 0) or 0),
        ])
    if len(daily_data) == 1:
        daily_data.append(["No data", "-", "-", "-"])
    t2 = Table(daily_data, colWidths=[1.5*inch, 1.2*inch, 1.2*inch, 1.2*inch])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#334155')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    story.append(t2)
    story.append(Spacer(1, 20))
    
    story.append(Paragraph("Incidents by Type", styles['Heading2']))
    inc_data = [["Incident Type", "Count"]]
    for inc in incident_types:
        inc_data.append([(inc.get('event_type') or '').replace('_', ' ').title(), str(inc.get('count', 0) or 0)])
    if len(inc_data) == 1:
        inc_data.append(["No incidents", "-"])
    t3 = Table(inc_data, colWidths=[3*inch, 1.5*inch])
    t3.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#334155')),
    ]))
    story.append(t3)
    story.append(Spacer(1, 20))
    
    story.append(Paragraph("Top User Performance", styles['Heading2']))
    perf_data = [["Rank", "User", "Exams", "Avg Score", "Warnings"]]
    for i, u in enumerate(user_performance, start=1):
        perf_data.append([
            str(i),
            u.get('full_name', '') or '',
            str(u.get('exams_taken', 0) or 0),
            str(round(u.get('avg_score') or 0, 1)),
            str(u.get('total_warnings') or 0),
        ])
    if len(perf_data) == 1:
        perf_data.append(["No data", "-", "-", "-", "-"])
    t4 = Table(perf_data, colWidths=[0.5*inch, 2*inch, 0.8*inch, 1*inch, 1*inch])
    t4.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#334155')),
    ]))
    story.append(t4)
    
    doc.build(story)
    buffer.seek(0)
    filename = f"analytics_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)


@app.route('/admin/exams')
@login_required
@admin_required
def admin_exams():
    """List and manage all exams"""
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT e.*, u.full_name as creator_name,
               COUNT(DISTINCT es.id) as attempt_count,
               COUNT(DISTINCT q.id) as question_count
        FROM exams e
        LEFT JOIN users u ON e.created_by = u.id
        LEFT JOIN exam_sessions es ON e.id = es.exam_id
        LEFT JOIN questions q ON e.id = q.exam_id
        GROUP BY e.id
        ORDER BY e.created_at DESC
    """)
    exams = cur.fetchall()
    cur.close()
    return render_template('admin/exams.html', exams=exams)


@app.route('/admin/exams/create', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_create_exam():
    """Create a new exam with questions"""
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        duration_minutes = request.form.get('duration_minutes', 60, type=int)
        total_marks = request.form.get('total_marks', 100, type=int)
        passing_marks = request.form.get('passing_marks', 40, type=int)
        is_active = 'is_active' in request.form

        if not title:
            flash('Exam title is required.', 'error')
            return render_template('admin/create_exam.html')

        if passing_marks > total_marks:
            flash('Passing marks cannot exceed total marks.', 'error')
            return render_template('admin/create_exam.html')

        questions_data = []
        q_index = 0
        while True:
            q_text = request.form.get(f'questions[{q_index}][text]')
            if q_text is None:
                break
            q_text = q_text.strip()
            opt_a = request.form.get(f'questions[{q_index}][option_a]', '').strip()
            opt_b = request.form.get(f'questions[{q_index}][option_b]', '').strip()
            opt_c = request.form.get(f'questions[{q_index}][option_c]', '').strip()
            opt_d = request.form.get(f'questions[{q_index}][option_d]', '').strip()
            correct = request.form.get(f'questions[{q_index}][correct]', 'A')
            marks = request.form.get(f'questions[{q_index}][marks]', 1, type=int)

            if q_text and opt_a and opt_b and opt_c and opt_d:
                questions_data.append({
                    'text': q_text, 'option_a': opt_a, 'option_b': opt_b,
                    'option_c': opt_c, 'option_d': opt_d,
                    'correct': correct, 'marks': marks
                })
            q_index += 1

        if not questions_data:
            flash('At least one complete question is required.', 'error')
            return render_template('admin/create_exam.html')

        cur = mysql.connection.cursor()
        cur.execute("""
            INSERT INTO exams (title, description, duration_minutes, total_marks, passing_marks, is_active, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (title, description, duration_minutes, total_marks, passing_marks, is_active, current_user.id))
        exam_id = cur.lastrowid

        for q in questions_data:
            cur.execute("""
                INSERT INTO questions (exam_id, question_text, option_a, option_b, option_c, option_d, correct_option, marks)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (exam_id, q['text'], q['option_a'], q['option_b'], q['option_c'], q['option_d'], q['correct'], q['marks']))

        mysql.connection.commit()
        cur.close()

        flash(f'Exam "{title}" created successfully with {len(questions_data)} questions.', 'success')
        return redirect(url_for('admin_exams'))

    return render_template('admin/create_exam.html')


@app.route('/admin/exams/<int:exam_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_exam(exam_id):
    """Edit an existing exam and its questions"""
    cur = mysql.connection.cursor()

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        duration_minutes = request.form.get('duration_minutes', 60, type=int)
        total_marks = request.form.get('total_marks', 100, type=int)
        passing_marks = request.form.get('passing_marks', 40, type=int)
        is_active = 'is_active' in request.form

        if not title:
            flash('Exam title is required.', 'error')
            return redirect(url_for('admin_edit_exam', exam_id=exam_id))

        if passing_marks > total_marks:
            flash('Passing marks cannot exceed total marks.', 'error')
            return redirect(url_for('admin_edit_exam', exam_id=exam_id))

        questions_data = []
        q_index = 0
        while True:
            q_text = request.form.get(f'questions[{q_index}][text]')
            if q_text is None:
                break
            q_text = q_text.strip()
            opt_a = request.form.get(f'questions[{q_index}][option_a]', '').strip()
            opt_b = request.form.get(f'questions[{q_index}][option_b]', '').strip()
            opt_c = request.form.get(f'questions[{q_index}][option_c]', '').strip()
            opt_d = request.form.get(f'questions[{q_index}][option_d]', '').strip()
            correct = request.form.get(f'questions[{q_index}][correct]', 'A')
            marks = request.form.get(f'questions[{q_index}][marks]', 1, type=int)

            if q_text and opt_a and opt_b and opt_c and opt_d:
                questions_data.append({
                    'text': q_text, 'option_a': opt_a, 'option_b': opt_b,
                    'option_c': opt_c, 'option_d': opt_d,
                    'correct': correct, 'marks': marks
                })
            q_index += 1

        if not questions_data:
            flash('At least one complete question is required.', 'error')
            return redirect(url_for('admin_edit_exam', exam_id=exam_id))

        cur.execute("""
            UPDATE exams SET title=%s, description=%s, duration_minutes=%s,
                   total_marks=%s, passing_marks=%s, is_active=%s
            WHERE id=%s
        """, (title, description, duration_minutes, total_marks, passing_marks, is_active, exam_id))

        cur.execute("DELETE FROM questions WHERE exam_id = %s", (exam_id,))

        for q in questions_data:
            cur.execute("""
                INSERT INTO questions (exam_id, question_text, option_a, option_b, option_c, option_d, correct_option, marks)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (exam_id, q['text'], q['option_a'], q['option_b'], q['option_c'], q['option_d'], q['correct'], q['marks']))

        mysql.connection.commit()
        cur.close()

        flash(f'Exam "{title}" updated successfully.', 'success')
        return redirect(url_for('admin_exams'))

    cur.execute("SELECT * FROM exams WHERE id = %s", (exam_id,))
    exam = cur.fetchone()

    if not exam:
        flash('Exam not found.', 'error')
        cur.close()
        return redirect(url_for('admin_exams'))

    cur.execute("SELECT * FROM questions WHERE exam_id = %s ORDER BY id", (exam_id,))
    questions = cur.fetchall()
    cur.close()

    return render_template('admin/edit_exam.html', exam=exam, questions=questions)


@app.route('/admin/exams/<int:exam_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_exam(exam_id):
    """Delete an exam and all its data"""
    cur = mysql.connection.cursor()

    cur.execute("SELECT title FROM exams WHERE id = %s", (exam_id,))
    exam = cur.fetchone()

    if not exam:
        flash('Exam not found.', 'error')
        cur.close()
        return redirect(url_for('admin_exams'))

    cur.execute("DELETE FROM exams WHERE id = %s", (exam_id,))
    mysql.connection.commit()
    cur.close()

    flash(f'Exam "{exam["title"]}" has been deleted.', 'success')
    return redirect(url_for('admin_exams'))


@app.route('/admin/exams/<int:exam_id>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_toggle_exam(exam_id):
    """Toggle exam active/inactive status"""
    cur = mysql.connection.cursor()

    cur.execute("SELECT id, is_active, title FROM exams WHERE id = %s", (exam_id,))
    exam = cur.fetchone()

    if not exam:
        flash('Exam not found.', 'error')
        cur.close()
        return redirect(url_for('admin_exams'))

    new_status = not exam['is_active']
    cur.execute("UPDATE exams SET is_active = %s WHERE id = %s", (new_status, exam_id))
    mysql.connection.commit()
    cur.close()

    status_text = 'activated' if new_status else 'deactivated'
    flash(f'Exam "{exam["title"]}" has been {status_text}.', 'success')
    return redirect(url_for('admin_exams'))


@app.route('/admin/session/<int:session_id>')
@login_required
@admin_required
def admin_view_session(session_id):
    """View detailed session information"""
    cur = mysql.connection.cursor()
    
    cur.execute("""
        SELECT es.*, u.username, u.full_name, u.email, e.title as exam_title, e.total_marks
        FROM exam_sessions es
        JOIN users u ON es.user_id = u.id
        JOIN exams e ON es.exam_id = e.id
        WHERE es.id = %s
    """, (session_id,))
    session_data = cur.fetchone()
    
    if not session_data:
        flash('Session not found.', 'error')
        cur.close()
        return redirect(url_for('admin_monitor'))
    
    # Get proctoring logs for this session
    cur.execute("""
        SELECT * FROM proctoring_logs
        WHERE session_id = %s
        ORDER BY created_at ASC
    """, (session_id,))
    logs = cur.fetchall()
    
    cur.close()
    
    return render_template('admin/session_detail.html', session=session_data, logs=logs)


@app.route('/api/proctoring_alerts')
@login_required
def get_proctoring_alerts():
    """Return recent warning/critical proctoring events for notifications.
    Admin: all such events in the last 10 minutes.
    User: only events for their own sessions in the last 10 minutes.
    Query param: since (ISO timestamp) to get only alerts after that time.
    """
    since_param = request.args.get('since')
    cur = mysql.connection.cursor()
    if current_user.is_admin():
        if since_param:
            cur.execute("""
                SELECT pl.id, pl.session_id, pl.user_id, pl.event_type, pl.severity, pl.description, pl.created_at,
                       u.username, u.full_name, e.title as exam_title
                FROM proctoring_logs pl
                JOIN users u ON pl.user_id = u.id
                JOIN exam_sessions es ON pl.session_id = es.id
                JOIN exams e ON es.exam_id = e.id
                WHERE pl.severity IN ('warning', 'critical') AND pl.created_at > %s
                ORDER BY pl.created_at DESC
                LIMIT 50
            """, (since_param,))
        else:
            cur.execute("""
                SELECT pl.id, pl.session_id, pl.user_id, pl.event_type, pl.severity, pl.description, pl.created_at,
                       u.username, u.full_name, e.title as exam_title
                FROM proctoring_logs pl
                JOIN users u ON pl.user_id = u.id
                JOIN exam_sessions es ON pl.session_id = es.id
                JOIN exams e ON es.exam_id = e.id
                WHERE pl.severity IN ('warning', 'critical')
                  AND pl.created_at >= DATE_SUB(NOW(), INTERVAL 10 MINUTE)
                ORDER BY pl.created_at DESC
                LIMIT 50
            """)
    else:
        if since_param:
            cur.execute("""
                SELECT pl.id, pl.session_id, pl.user_id, pl.event_type, pl.severity, pl.description, pl.created_at,
                       u.username, u.full_name, e.title as exam_title
                FROM proctoring_logs pl
                JOIN users u ON pl.user_id = u.id
                JOIN exam_sessions es ON pl.session_id = es.id
                JOIN exams e ON es.exam_id = e.id
                WHERE pl.user_id = %s AND pl.severity IN ('warning', 'critical') AND pl.created_at > %s
                ORDER BY pl.created_at DESC
                LIMIT 50
            """, (current_user.id, since_param))
        else:
            cur.execute("""
                SELECT pl.id, pl.session_id, pl.user_id, pl.event_type, pl.severity, pl.description, pl.created_at,
                       u.username, u.full_name, e.title as exam_title
                FROM proctoring_logs pl
                JOIN users u ON pl.user_id = u.id
                JOIN exam_sessions es ON pl.session_id = es.id
                JOIN exams e ON es.exam_id = e.id
                WHERE pl.user_id = %s AND pl.severity IN ('warning', 'critical')
                  AND pl.created_at >= DATE_SUB(NOW(), INTERVAL 10 MINUTE)
                ORDER BY pl.created_at DESC
                LIMIT 50
            """, (current_user.id,))
    rows = cur.fetchall()
    cur.close()
    alerts = []
    for r in rows:
        alerts.append({
            'id': r['id'],
            'session_id': r['session_id'],
            'user_id': r['user_id'],
            'event_type': r['event_type'],
            'severity': r['severity'],
            'description': r['description'] or '',
            'created_at': r['created_at'].strftime('%Y-%m-%dT%H:%M:%S') if r['created_at'] else None,
            'username': r['username'],
            'full_name': r['full_name'],
            'exam_title': r['exam_title'],
        })
    return jsonify(alerts)


@app.route('/api/active_sessions')
@login_required
@admin_required
def get_active_sessions():
    """API endpoint for live monitoring"""
    cur = mysql.connection.cursor()
    
    cur.execute("""
        SELECT es.id, es.warning_count, es.face_verified, es.start_time,
               u.username, u.full_name, e.title as exam_title,
               (SELECT COUNT(*) FROM proctoring_logs WHERE session_id = es.id AND severity = 'critical') as critical_count
        FROM exam_sessions es
        JOIN users u ON es.user_id = u.id
        JOIN exams e ON es.exam_id = e.id
        WHERE es.status = 'in_progress'
        ORDER BY es.start_time DESC
    """)
    sessions = cur.fetchall()
    cur.close()
    
    # Convert datetime objects to strings
    for session in sessions:
        if session['start_time']:
            session['start_time'] = session['start_time'].strftime('%Y-%m-%d %H:%M:%S')
    
    return jsonify(sessions)


# =============================================
# Error Handlers
# =============================================
@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404


@app.errorhandler(500)
def internal_error(error):
    return render_template('errors/500.html'), 500


# =============================================
# Run Application
# =============================================
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
